"""
CIS School System - Excel Upload Routes
=========================================
Allows admin to upload Excel files to import:
  - Students list
  - Comments
  - Marks (Entry, Mid Term, End Term)
"""

import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from ..models import db, Grade, Stream, Student, Term, Assessment, Mark, ReportCard
from ..grading import combine_split_subject, assign_performance_level
from .auth import login_required, role_required

upload_bp = Blueprint("upload", __name__)


def title_case(name):
    return " ".join(w.capitalize() for w in str(name).strip().split())


def find_or_create_student(name, grade, stream_name):
    """Find existing student or create new one."""
    stream = Stream.query.filter_by(grade_id=grade.id, name=stream_name.upper()).first()
    if not stream:
        stream = Stream.query.filter_by(grade_id=grade.id).first()

    # Try exact match first
    student = Student.query.filter_by(full_name=name, grade_id=grade.id).first()
    if student:
        return student

    # Try first name match
    first = name.split()[0] if name.split() else name
    student = Student.query.filter(
        Student.grade_id == grade.id,
        Student.full_name.ilike(f"{first}%")
    ).first()
    if student:
        return student

    # Create new
    existing_count = Student.query.filter_by(grade_id=grade.id).count()
    grade_code = grade.name.replace("Grade ", "G").replace(" ", "")
    adm = f"CIS-{grade_code}{stream_name[:1] if stream_name else 'R'}-{existing_count+1:03d}"

    student = Student(
        full_name=name,
        admission_no=adm,
        grade_id=grade.id,
        stream_id=stream.id if stream else None,
        is_active=True,
    )
    db.session.add(student)
    db.session.flush()
    return student


@upload_bp.route("/")
@login_required
@role_required("admin", "principal")
def index():
    active_term = Term.query.filter_by(is_active=True).first()
    assessments = Assessment.query.filter_by(term_id=active_term.id).all() if active_term else []
    student_count = Student.query.filter_by(is_active=True).count()
    return render_template("upload/index.html",
                           active_term=active_term,
                           assessments=assessments,
                           student_count=student_count)


@upload_bp.route("/students", methods=["POST"])
@login_required
@role_required("admin", "principal")
def upload_students():
    """Upload students + comments from Excel (COMMENTS.xlsx format)."""
    import openpyxl
    from io import BytesIO

    file = request.files.get("excel_file")
    if not file:
        flash("No file uploaded.", "danger")
        return redirect(url_for("upload.index"))

    active_term = Term.query.filter_by(is_active=True).first()

    GRADE_MAP = {
        "PP2": "PP2", "PP1": "PP1",
        "GRADE 1": "Grade 1", "GRADE 2": "Grade 2", "GRADE 3": "Grade 3",
        "GRADE 4": "Grade 4", "GRADE 5": "Grade 5", "GRADE 6": "Grade 6",
        "GRADE 7": "Grade 7", "GRADE 8": "Grade 8", "GRADE 9": "Grade 9",
    }

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        added = 0
        skipped = 0
        comments_loaded = 0

        for sn in wb.sheetnames:
            grade_name = GRADE_MAP.get(sn.strip().upper())
            if not grade_name:
                continue

            grade = Grade.query.filter_by(name=grade_name).first()
            if not grade:
                continue

            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))

            # Find header
            header_idx = 0
            header = []
            for i, row in enumerate(rows):
                cells = [str(c).strip().upper() if c else "" for c in row]
                if "NAME" in cells or "LEARNER'S NAME" in cells:
                    header = cells
                    header_idx = i
                    break

            if not header:
                continue

            def ci(opts):
                for o in opts:
                    for i, h in enumerate(header):
                        if o in h:
                            return i
                return None

            nc = ci(["LEARNER'S NAME", "NAME"])
            sc = ci(["STREAM"])
            pc = ci(["PERFORMANCE"])
            cc = ci(["COMPETENCIES"])
            vc = ci(["VALUES"])

            def g(row, idx):
                if idx is not None and idx < len(row) and row[idx]:
                    return str(row[idx]).strip()
                return ""

            for row in rows[header_idx + 1:]:
                if not row or nc is None or not row[nc]:
                    continue
                name = title_case(str(row[nc]))
                if not name or name.upper() in ("NAME",):
                    continue

                stream_name = g(row, sc).upper() or "RED"
                student = find_or_create_student(name, grade, stream_name)

                if student.id is None:
                    added += 1
                else:
                    skipped += 1

                # Load comments
                if active_term and (pc or vc):
                    perf = g(row, pc)
                    comp = g(row, cc)
                    vals = g(row, vc)
                    if perf or vals:
                        rc = ReportCard.query.filter_by(
                            student_id=student.id, term_id=active_term.id
                        ).first()
                        if not rc:
                            rc = ReportCard(student_id=student.id, term_id=active_term.id)
                            db.session.add(rc)
                        rc.comment_performance = perf
                        rc.comment_competencies = comp
                        rc.comment_values = vals
                        rc.status = "pending_approval"
                        comments_loaded += 1

            db.session.commit()

        wb.close()
        flash(f"✅ Import complete — {added} students added, {comments_loaded} comments loaded.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    return redirect(url_for("upload.index"))


@upload_bp.route("/marks", methods=["POST"])
@login_required
@role_required("admin", "principal")
def upload_marks():
    """Upload marks from Entry/Mid/End Term Excel files."""
    import openpyxl
    from io import BytesIO

    file = request.files.get("excel_file")
    assessment_number = int(request.form.get("assessment_number", 2))

    if not file:
        flash("No file uploaded.", "danger")
        return redirect(url_for("upload.index"))

    active_term = Term.query.filter_by(is_active=True).first()
    if not active_term:
        flash("No active term.", "warning")
        return redirect(url_for("upload.index"))

    assessment = Assessment.query.filter_by(
        term_id=active_term.id, number=assessment_number
    ).first()
    if not assessment:
        flash("Assessment not found.", "warning")
        return redirect(url_for("upload.index"))

    GRADE_MAP = {
        "GRADE 9": "Grade 9", "GRADE 8": "Grade 8", "GRADE 7": "Grade 7",
        "GRADE 6": "Grade 6", "GRADE 5": "Grade 5", "GRADE 4": "Grade 4",
        "GRADE 4 ": "Grade 4", "GRADE 3": "Grade 3", "GRADE 2": "Grade 2",
        "GRADE 1": "Grade 1", "PP2 RED": "PP2", "PP2 YELLOW": "PP2",
        "PP1 RED": "PP1", "PP1 YELLOW": "PP1", "PP2": "PP2", "PP1": "PP1",
        "GRADE 3 ": "Grade 3",
    }

    SUBJECTS_LOWER = ["Mathematics", "English", "Kiswahili", "Environmental Activities",
                      "Christian Religious Education", "Creative Arts"]
    SUBJECTS_UPPER = ["Mathematics", "English", "Kiswahili", "Science & Technology",
                      "Social Studies", "Christian Religious Education", "Creative Arts", "Agriculture"]
    SUBJECTS_JUNIOR = ["Mathematics", "English", "Kiswahili", "Integrated Science",
                       "Social Studies", "Christian Religious Education", "Creative Arts",
                       "Agriculture", "Pre-Technical Studies"]
    SUBJECTS_PRE = ["Mathematics Activities", "Language Activities", "Environmental Activities",
                    "Christian Religious Education", "Creative Arts"]

    try:
        wb = openpyxl.load_workbook(BytesIO(file.read()), read_only=True, data_only=True)
        total_saved = 0

        for sn in wb.sheetnames:
            grade_name = GRADE_MAP.get(sn.strip())
            if not grade_name:
                continue

            grade = Grade.query.filter_by(name=grade_name).first()
            if not grade:
                continue

            level = grade.level_group
            if level == "junior":
                subjects = SUBJECTS_JUNIOR
            elif level == "upper_primary":
                subjects = SUBJECTS_UPPER
            elif level == "lower_primary":
                subjects = SUBJECTS_LOWER
            else:
                subjects = SUBJECTS_PRE

            SPLIT = {}
            if level == "upper_primary":
                SPLIT = {"English": (40, 10), "Kiswahili": (40, 10)}
            elif level == "junior":
                SPLIT = {"English": (50, 50), "Kiswahili": (50, 50)}

            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))

            header_idx = 0
            header = []
            for i, row in enumerate(rows):
                cells = [str(c).strip().upper() if c else "" for c in row]
                if "NAME" in cells:
                    header = cells
                    header_idx = i
                    break
            if not header:
                continue

            ni = next((i for i, h in enumerate(header) if "NAME" in h), 1)
            si = next((i for i, h in enumerate(header) if "STREAM" in h), None)

            for row in rows[header_idx + 1:]:
                if not row or not row[ni]:
                    continue
                name = title_case(str(row[ni]))
                if not name or name[0].isdigit():
                    continue

                stream_val = ""
                if si and si < len(row) and row[si]:
                    stream_val = str(row[si]).strip().upper()
                elif "RED" in sn.upper():
                    stream_val = "RED"
                elif "YELLOW" in sn.upper():
                    stream_val = "YELLOW"

                student = Student.query.filter_by(full_name=name, grade_id=grade.id).first()
                if not student:
                    first = name.split()[0]
                    student = Student.query.filter(
                        Student.grade_id == grade.id,
                        Student.full_name.ilike(f"{first}%")
                    ).first()
                if not student:
                    continue

                def get_val(col_idx):
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            return float(row[col_idx])
                        except:
                            return None
                    return None

                # Map columns based on header
                offset = ni + 1
                if si and si == offset:
                    offset += 1

                for subj in subjects:
                    if subj in SPLIT:
                        p1 = get_val(offset)
                        p2 = get_val(offset + 1)
                        offset += 3  # p1, p2, total/pct
                        p1max, p2max = SPLIT[subj]
                        combined = combine_split_subject(p1, p2, p1max, p2max) if p1 and p2 else None
                        effective = combined
                        code, label = assign_performance_level(effective, grade_name) if effective else (None, None)
                        mark = Mark.query.filter_by(
                            student_id=student.id, assessment_id=assessment.id, subject=subj
                        ).first()
                        if not mark:
                            mark = Mark(student_id=student.id, assessment_id=assessment.id, subject=subj)
                            db.session.add(mark)
                        # Save as single score — use combined % or p1 directly
                        max_s = 100 if get_grade_level(grade_name) == "junior" else 30
                        raw_eff = float(combined) if combined else (float(p1) if p1 else None)
                        effective = min(int(round(raw_eff)), max_s) if raw_eff is not None else None
                        code, label = assign_performance_level(effective, grade_name) if effective else (None, None)
                        mark.score = effective
                        mark.paper1_score = None
                        mark.paper2_score = None
                        mark.combined_score = None
                        mark.grade_code = code
                        mark.grade_label = label
                    else:
                        sc_val = get_val(offset)
                        offset += 1
                        code, label = assign_performance_level(sc_val, grade_name) if sc_val else (None, None)
                        mark = Mark.query.filter_by(
                            student_id=student.id, assessment_id=assessment.id, subject=subj
                        ).first()
                        if not mark:
                            mark = Mark(student_id=student.id, assessment_id=assessment.id, subject=subj)
                            db.session.add(mark)
                        mark.score = sc_val
                        mark.paper1_score = None
                        mark.paper2_score = None
                        mark.combined_score = None
                        mark.grade_code = code
                        mark.grade_label = label

                total_saved += 1

            db.session.commit()

        wb.close()
        ass_name = {1: "Entry Assessment", 2: "Mid Term", 3: "End Term"}.get(assessment_number)
        flash(f"✅ {ass_name} marks loaded — {total_saved} students updated.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error: {str(e)}", "danger")

    return redirect(url_for("upload.index"))
